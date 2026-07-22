"""
extract_actividad_ipi.py
========================
Lee IPI - Todos.xlsx (hoja 'Hoja1') y genera
assets/data/actividad_ipi.js para el dashboard EcoGo.

Series: EMAE, IPI Manufacturero, ISAC, IPI Minero, ISSP
Base: nov-23 = 100 · Serie sin estacionalidad (columnas G-K)

Uso directo:    python extract_actividad_ipi.py
Desde refresh:  from extract_actividad_ipi import run_extraction
"""

import os, re, json
from datetime import datetime

EXCEL_NAME  = "IPI - Todos.xlsx"
SHEET_NAME  = "Hoja1"
SERIES_COLS = {          # columna (1-based) → nombre
    7:  "EMAE",
    8:  "IPI Manufacturero",
    9:  "ISAC",
    10: "IPI Minero",
    11: "ISSP",
}
DATE_COL = 1  # columna A


def run_extraction(excel_path: str, dashboard_dir: str) -> dict:
    """
    Lee IPI - Todos.xlsx y escribe assets/data/actividad_ipi.js.
    Retorna {'ok': bool, 'msg': str}.
    """
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "msg": "openpyxl no instalado (pip install openpyxl)"}

    if not os.path.exists(excel_path):
        return {"ok": False, "msg": f"No se encontró: {excel_path}"}

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
    except Exception as e:
        return {"ok": False, "msg": f"Error abriendo Excel: {e}"}

    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return {"ok": False, "msg": f"Hoja '{SHEET_NAME}' no encontrada en {os.path.basename(excel_path)}"}

    ws = wb[SHEET_NAME]

    dates = []
    series = {name: [] for name in SERIES_COLS.values()}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date_val = row[DATE_COL - 1]
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            d = date_val
        else:
            try:
                d = datetime(int(date_val), 1, 1)
            except Exception:
                continue

        dates.append(d.strftime("%Y-%m"))
        for col_idx, name in SERIES_COLS.items():
            val = row[col_idx - 1]
            if val is None or (isinstance(val, str) and val.startswith("#")):
                series[name].append(None)
            else:
                try:
                    series[name].append(round(float(val), 4))
                except (TypeError, ValueError):
                    series[name].append(None)

    wb.close()

    if not dates:
        return {"ok": False, "msg": "No se encontraron datos en Hoja1"}

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    data = {
        "generated": ts,
        "source":    f"{EXCEL_NAME} · {SHEET_NAME}",
        "base":      "nov-23=100 · Serie sin estacionalidad",
        "dates":     dates,
        "series":    series,
    }

    out_path = os.path.join(dashboard_dir, "assets", "data", "actividad_ipi.js")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"// Indicadores de actividad - regenerado el {ts}\n")
        f.write("window.ACTIVIDAD_IPI = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n")

    sz = os.path.getsize(out_path)
    n_series = len(SERIES_COLS)
    return {"ok": True, "msg": f"{sz:,} bytes · {n_series} series · {dates[0]}–{dates[-1]}"}


# ── Ejecución directa ──────────────────────────────────────────
if __name__ == "__main__":
    import sys, platform

    script_dir = os.path.dirname(os.path.abspath(__file__))

    if platform.system() == "Windows":
        base_excel = r"C:\Users\fscalise\OneDrive - ECOGO S.A\BD\Actividad"
    else:
        base_excel = "/sessions/sleepy-upbeat-cannon/mnt/BD/Actividad"

    excel = os.path.join(base_excel, EXCEL_NAME)
    print(f"Leyendo: {excel}")
    result = run_extraction(excel, script_dir)
    if result["ok"]:
        print(f"  [OK] Actividad IPI — {result['msg']}")
        sys.exit(0)
    else:
        print(f"  [FAIL] {result['msg']}")
        sys.exit(1)
