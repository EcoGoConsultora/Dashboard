"""
extract_series_largas.py
========================
Lee Anexo.xlsx (hoja 'Monetaria - SPNF-Deuda-ML') y genera
assets/data/series-largas.js para el dashboard EcoGo.

Reglas:
- Solo filas VISIBLES (respeta filas ocultas del Excel)
- Valores cacheados de fórmulas (data_only=True)
- Los Excel NO se modifican nunca
- Actualiza también la descripción en layout.js y layout-clientes.js

Uso directo:    python extract_series_largas.py
Desde refresh:  from extract_series_largas import run_extraction
"""

import os, re, json
from datetime import datetime

SHEET_NAME = "Monetaria - SPNF-Deuda-ML"

# Estructura fija de la hoja: (label, range de filas 1-based)
# Si en el futuro se agregan/mueven filas, ajustar aquí.
CAT_DEFS = [
    ("Monetarias y Financieras",              range(3, 24)),
    ("Sector Público Nacional No Financiero", range(25, 35)),
    ("Deuda",                                 range(36, 43)),
    ("Mercado Laboral",                       range(44, 47)),
    ("Actividad (INDEC) Base 2004=100",       range(48, 59)),
    ("Precios",                               range(60, 77)),
    ("Sector Externo",                        range(78, 86)),
]


def _slugify(s):
    s = str(s).lower().strip()
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"[\s_]+", "-", s)
    return re.sub(r"-+", "-", s).strip("-")[:60]


def run_extraction(anexo_path: str, dashboard_dir: str) -> dict:
    """
    Lee Anexo.xlsx y escribe assets/data/series-largas.js.
    Retorna {'ok': bool, 'msg': str, 'last_year': int|None}.
    """
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "msg": "openpyxl no instalado (pip install openpyxl)", "last_year": None}

    if not os.path.exists(anexo_path):
        return {"ok": False, "msg": f"No se encontró: {anexo_path}", "last_year": None}

    try:
        # read_only=False para acceder a row_dimensions (filas ocultas)
        wb = openpyxl.load_workbook(anexo_path, read_only=False, data_only=True)
    except Exception as e:
        return {"ok": False, "msg": f"Error abriendo Excel: {e}", "last_year": None}

    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return {"ok": False, "msg": f"Hoja '{SHEET_NAME}' no encontrada en {os.path.basename(anexo_path)}", "last_year": None}

    ws = wb[SHEET_NAME]

    # Años en fila 1 (columna B en adelante)
    years, year_cols = [], []
    for cell in ws[1]:
        if cell.column == 1:
            continue
        try:
            y = int(cell.value)
            if 1985 <= y <= 2050:
                years.append(y)
                year_cols.append(cell.column)
        except (TypeError, ValueError):
            pass

    # Filas ocultas
    hidden = {idx for idx, rd in ws.row_dimensions.items() if rd.hidden}

    categories = []
    for cat_label, row_range in CAT_DEFS:
        variables = []
        for r in row_range:
            if r in hidden:
                continue
            name = ws.cell(row=r, column=1).value
            if not name or not str(name).strip():
                continue
            yr_vals = [ws.cell(row=r, column=c).value for c in year_cols]
            # Limpiar errores Excel (#REF!, #VALUE!, etc.)
            clean = [
                None if (v is None or (isinstance(v, str) and v.startswith("#")))
                else v
                for v in yr_vals
            ]
            if all(v is None for v in clean):
                continue
            variables.append({
                "id":     _slugify(name),
                "label":  str(name).strip(),
                "values": clean,
            })
        if variables:
            categories.append({
                "id":        _slugify(cat_label),
                "label":     cat_label,
                "variables": variables,
            })

    wb.close()

    last_year  = years[-1] if years else None
    first_year = years[0]  if years else None
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    data = {
        "generated":  ts,
        "source":     f"Anexo.xlsx · {SHEET_NAME} (filas visibles)",
        "years":      years,
        "categories": categories,
    }

    # Escribir series-largas.js
    out_path = os.path.join(dashboard_dir, "assets", "data", "series-largas.js")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"// Series Largas - regenerado el {ts}\n")
        f.write("window.SERIES_DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n")
    sz = os.path.getsize(out_path)

    # Actualizar año en la descripción de layout.js y layout-clientes.js
    if first_year and last_year:
        for fname in ("layout.js", "layout-clientes.js"):
            lpath = os.path.join(dashboard_dir, "assets", "js", fname)
            if not os.path.exists(lpath):
                continue
            with open(lpath, "r", encoding="utf-8") as f:
                txt = f.read()
            new_txt = re.sub(
                r"(Variables macroecon[oó]micas hist[oó]ricas )\d{4}[–-]\d{4}",
                rf"\g<1>{first_year}–{last_year}",
                txt,
            )
            if new_txt != txt:
                with open(lpath, "w", encoding="utf-8") as f:
                    f.write(new_txt)

    total_vars = sum(len(c["variables"]) for c in categories)
    return {
        "ok":        True,
        "msg":       f"{sz:,} bytes · {len(categories)} categorías · {total_vars} variables · hasta {last_year}",
        "last_year": last_year,
    }


# ── Ejecución directa ──────────────────────────────────────────
if __name__ == "__main__":
    import sys, platform

    script_dir = os.path.dirname(os.path.abspath(__file__))

    if platform.system() == "Windows":
        base_excel = r"C:\Users\fscalise\OneDrive - ECOGO S.A\BD"
    else:
        base_excel = os.path.dirname(os.path.dirname(script_dir))

    anexo = os.path.join(base_excel, "03 Informes y Anexos", "Cuadros y Anexos", "Anexos nuevos", "Anexo.xlsx")
    print(f"Leyendo: {anexo}")
    result = run_extraction(anexo, script_dir)
    if result["ok"]:
        print(f"  [OK] Series Largas — {result['msg']}")
        sys.exit(0)
    else:
        print(f"  [FAIL] {result['msg']}")
        sys.exit(1)
